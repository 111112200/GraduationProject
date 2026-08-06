import io
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.schemas.check import CreateCheckTaskRequest
from app.models import CheckTask, CheckResultSummary, CheckResultDetail, Experiment, Report, User
from app.api.deps import get_current_user
from app.services.check_validation import validate_check_reports

router = APIRouter()


@router.get("")
async def api_list_checks(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    tasks = db.query(CheckTask).filter(CheckTask.user_id == current_user.id).order_by(CheckTask.created_at.desc()).limit(50).all()
    return {
        "tasks": [
            {
                "taskId": t.id,
                "name": t.name,
                "status": t.status,
                "createdAt": t.created_at.isoformat() if t.created_at else None,
            }
            for t in tasks
        ]
    }


def run_check_task(task_id: int):
    from app.core.database import SessionLocal
    from app.services.check_task_service import execute_check_task
    db = SessionLocal()
    try:
        execute_check_task(db, task_id)
    finally:
        db.close()


@router.post("")
async def api_create_check(
    body: CreateCheckTaskRequest,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    experiment = db.query(Experiment).filter(
        Experiment.id == body.experimentId,
        Experiment.user_id == current_user.id,
    ).first()
    if not experiment:
        raise HTTPException(
            status_code=404,
            detail="实验不存在或无权使用",
        )

    # 校验所有报告必须属于当前用户，再创建任务，避免跨用户关联写入数据库。
    reports = db.query(Report).filter(
        Report.id.in_(body.reportIds),
        Report.user_id == current_user.id,
    ).all()
    if len(reports) != len(body.reportIds):
        raise HTTPException(
            status_code=400,
            detail="部分报告不存在或不属于当前用户",
        )
    try:
        validate_check_reports(reports, experiment.id, body.mode)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    task = CheckTask(
        user_id=current_user.id,
        name=body.name,
        experiment_id=experiment.id,
        mode=body.mode,
        high_risk_threshold=body.highRiskThreshold,
        similar_threshold=body.similarThreshold,
    )
    db.add(task)
    for r in reports:
        task.reports.append(r)
    db.commit()

    background_tasks.add_task(run_check_task, task.id)
    return {"taskId": task.id, "status": "PENDING"}


@router.get("/{task_id}/export")
async def api_export_check_result(
    task_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """导出查重任务结果为 Excel 文件"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    task = db.query(CheckTask).filter(CheckTask.id == task_id, CheckTask.user_id == current_user.id).first()
    if not task:
        raise HTTPException(404, "任务不存在")
    if task.status != "COMPLETED":
        raise HTTPException(400, "任务尚未完成，无法导出")

    # 查询所有 summary + report 信息
    summaries = db.query(CheckResultSummary, Report).join(
        Report, CheckResultSummary.report_id == Report.id
    ).filter(CheckResultSummary.check_task_id == task_id).all()

    if not summaries:
        raise HTTPException(400, "暂无查重数据可导出")

    # ============ 样式定义 ============
    risk_map = {"HIGH": "高风险", "MEDIUM": "中风险", "LOW": "低风险"}
    mode_map = {"IN_CLASS": "班内互查", "HISTORY": "历史底库"}
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="3F51B5", end_color="3F51B5", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_align = Alignment(vertical="top", wrap_text=True)
    high_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    medium_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    low_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    risk_fill_map = {"HIGH": high_fill, "MEDIUM": medium_fill, "LOW": low_fill}

    def _auto_width(ws_target, max_width=50, min_width=8):
        """自适应列宽，中文字符按2个宽度计算"""
        for col in ws_target.columns:
            col_max = min_width
            for cell in col:
                try:
                    val = str(cell.value or "")
                    # 截取前100字符估算宽度，避免超长文本影响性能
                    sample = val[:100]
                    cjk = sum(1 for ch in sample if '\u4e00' <= ch <= '\u9fff')
                    w = len(sample) + cjk
                    if w > col_max:
                        col_max = w
                except Exception:
                    pass
            ws_target.column_dimensions[col[0].column_letter].width = min(col_max + 3, max_width)

    def _write_header(ws_target, headers, row=1):
        for col_idx, h in enumerate(headers, start=1):
            cell = ws_target.cell(row=row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    # ============ 预加载所有 detail 及 target 报告信息 ============
    summary_ids = [s.id for s, _ in summaries]
    all_details = db.query(CheckResultDetail).filter(
        CheckResultDetail.summary_id.in_(summary_ids)
    ).order_by(CheckResultDetail.similarity.desc()).all()

    target_report_ids = set()
    for d in all_details:
        target_report_ids.add(d.target_report_id)

    # 批量查询 target 报告
    target_reports_map = {}
    if target_report_ids:
        target_reports = db.query(Report).filter(
            Report.id.in_(target_report_ids),
            Report.user_id == current_user.id,
        ).all()
        for tr in target_reports:
            target_reports_map[tr.id] = tr
    all_details = [d for d in all_details if d.target_report_id in target_reports_map]

    # 按 summary_id 分组，只保留当前用户有权访问的目标报告详情
    detail_map = {}
    for d in all_details:
        detail_map.setdefault(d.summary_id, []).append(d)

    # ============ Sheet 1: 查重汇总 ============
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "查重汇总"

    s1_headers = ["序号", "学号", "姓名", "实验报告文件名", "总体相似度", "风险等级",
                  "最高片段相似度", "最相似对象", "相似片段数", "查重时间"]
    _write_header(ws1, s1_headers)

    for idx, (summary, report) in enumerate(summaries, start=1):
        details = detail_map.get(summary.id, [])
        max_sim = 0.0
        most_similar_target = "-"

        if details:
            top_detail = max(details, key=lambda d: d.similarity)
            max_sim = top_detail.similarity
            tr = target_reports_map.get(top_detail.target_report_id)
            if tr:
                parts = [p for p in [tr.student_name, tr.file_name] if p]
                most_similar_target = " - ".join(parts) if parts else "-"

        row_data = [
            idx,
            report.student_id or "-",
            report.student_name or "-",
            report.file_name or "-",
            f"{summary.overall_score * 100:.1f}%",
            risk_map.get(summary.risk_level, summary.risk_level),
            f"{max_sim * 100:.1f}%",
            most_similar_target,
            len(details),
            summary.created_at.strftime("%Y-%m-%d %H:%M:%S") if summary.created_at else "-",
        ]

        row_idx = idx + 1  # +1 因为表头占了第 1 行
        bg_fill = risk_fill_map.get(summary.risk_level)
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = center_align
            # 只给“总体相似度”列（第5列）上色
            if col_idx == 5 and bg_fill:
                cell.fill = bg_fill

    _auto_width(ws1)

    # ============ Sheet 2: 相似详情 ============
    ws2 = wb.create_sheet(title="相似详情")

    s2_headers = ["序号", "学号", "姓名", "报告文件名", "原始文本",
                  "相似文本", "相似度", "相似报告作者", "相似报告文件名", "来源"]
    _write_header(ws2, s2_headers)

    detail_row = 2
    report_idx = 0
    for summary, report in summaries:
        details = detail_map.get(summary.id, [])
        report_idx += 1

        if not details:
            row_data = [
                report_idx,
                report.student_id or "-",
                report.student_name or "-",
                report.file_name or "-",
                "（无相似片段）",
                "-",
                "-",
                "-",
                "-",
                "-",
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws2.cell(row=detail_row, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = wrap_align if col_idx in (5, 6) else center_align
            detail_row += 1
            continue

        group_start = detail_row
        for d in details:
            tr = target_reports_map.get(d.target_report_id)
            target_name = tr.student_name if tr and tr.student_name else "-"
            target_file = tr.file_name if tr and tr.file_name else "-"
            source_label = mode_map.get(d.mode, d.mode)

            row_data = [
                report_idx,
                report.student_id or "-",
                report.student_name or "-",
                report.file_name or "-",
                (d.source_text or "")[:500],
                (d.target_text or "")[:500],
                f"{d.similarity * 100:.1f}%",
                target_name,
                target_file,
                source_label,
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws2.cell(row=detail_row, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = wrap_align if col_idx in (5, 6) else center_align
                # 只给原始文本（第5列）和相似文本（第6列）上色
                if col_idx in (5, 6):
                    if d.similarity >= 0.8:
                        cell.fill = high_fill
                    elif d.similarity >= 0.5:
                        cell.fill = medium_fill

            detail_row += 1

        # 合并同一份报告的 序号/学号/姓名/文件名 列
        group_end = detail_row - 1
        if group_end > group_start:
            for merge_col in [1, 2, 3, 4]:
                ws2.merge_cells(
                    start_row=group_start, start_column=merge_col,
                    end_row=group_end, end_column=merge_col,
                )
                ws2.cell(row=group_start, column=merge_col).alignment = center_align

    # 设置详情表列宽
    detail_col_widths = {
        "A": 6,   # 序号
        "B": 12,  # 学号
        "C": 10,  # 姓名
        "D": 20,  # 报告文件名
        "E": 45,  # 原始文本
        "F": 45,  # 相似文本
        "G": 10,  # 相似度
        "H": 12,  # 相似报告作者
        "I": 20,  # 相似报告文件名
        "J": 12,  # 来源
    }
    for col_letter, width in detail_col_widths.items():
        ws2.column_dimensions[col_letter].width = width

    # 输出文件流
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{task.name}_查重结果报告.xlsx"
    encoded_filename = quote(filename)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
        },
    )


@router.get("/{task_id}")
async def api_get_check_task(
    task_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    task = db.query(CheckTask).filter(CheckTask.id == task_id, CheckTask.user_id == current_user.id).first()
    if not task:
        raise HTTPException(404, "任务不存在")
    summaries = db.query(CheckResultSummary, Report).join(
        Report, CheckResultSummary.report_id == Report.id
    ).filter(CheckResultSummary.check_task_id == task_id).all()
    results = [
        {
            "reportId": s.report_id,
            "studentName": r.student_name,
            "fileName": r.file_name,
            "overallScore": s.overall_score,
            "riskLevel": s.risk_level,
        }
        for s, r in summaries
    ]
    return {
        "taskId": task.id,
        "name": task.name,
        "status": task.status,
        "createdAt": task.created_at.isoformat() if task.created_at else None,
        "results": results,
    }


@router.delete("/{task_id}")
async def api_delete_check_task(
    task_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    task = db.query(CheckTask).filter(CheckTask.id == task_id, CheckTask.user_id == current_user.id).first()
    if not task:
        raise HTTPException(404, "任务不存在")
    # 清理 ChromaDB 中的临时向量索引
    from app.services.vector_store_service import delete_task_collection
    delete_task_collection(task_id)
    
    # 清理数据库中关联的查重结果
    summaries = db.query(CheckResultSummary).filter(CheckResultSummary.check_task_id == task_id).all()
    summary_ids = [s.id for s in summaries]
    if summary_ids:
        db.query(CheckResultDetail).filter(CheckResultDetail.summary_id.in_(summary_ids)).delete(synchronize_session=False)
        db.query(CheckResultSummary).filter(CheckResultSummary.check_task_id == task_id).delete(synchronize_session=False)

    db.delete(task)
    db.commit()
    return {"success": True}

